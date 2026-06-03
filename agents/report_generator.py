import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment
from datetime import datetime
from core.state import AgentState


# ── Styles ────────────────────────────────────────────────────────────────────

def header_fill():  return PatternFill("solid", fgColor="1F3864")
def sub_fill():     return PatternFill("solid", fgColor="2E75B6")
def green_fill():   return PatternFill("solid", fgColor="375623")
def med_fill():     return PatternFill("solid", fgColor="C6EFCE")
def low_fill():     return PatternFill("solid", fgColor="FFF2CC")
def red_fill():     return PatternFill("solid", fgColor="FF9999")
def weak_fill():    return PatternFill("solid", fgColor="FFD7D7")
def good_fill():    return PatternFill("solid", fgColor="D6F0D6")

def thin_border():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)

def hdr(cell, text, size=11):
    cell.value = text
    cell.font = Font(bold=True, color="FFFFFF", size=size)
    cell.fill = header_fill()
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = thin_border()

def sub(cell, text):
    cell.value = text
    cell.font = Font(bold=True, color="FFFFFF", size=10)
    cell.fill = sub_fill()
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = thin_border()

def dat(cell, value, center=False, wrap=False):
    cell.value = value
    cell.border = thin_border()
    cell.alignment = Alignment(
        horizontal="center" if center else "left",
        vertical="center",
        wrap_text=wrap
    )


# ── Sheet 6A — Competency Matrix ──────────────────────────────────────────────

def write_6A(ws, state: AgentState):
    ws.title = "6A - Competency Matrix"
    cos = state.cos
    pis = state.performance_indicators
    co_ids = [co.co_id for co in cos]

    # Title
    num_cols = len(co_ids) + 3
    ws.merge_cells(f"A1:{get_column_letter(num_cols)}1")
    hdr(ws["A1"], "COMPETENCY MATRIX (Performance Indicator to CO Mapping)", size=13)
    ws.row_dimensions[1].height = 30

    # Column headers
    sub(ws.cell(2, 1), "PO")
    sub(ws.cell(2, 2), "Competency")
    sub(ws.cell(2, 3), "Performance Indicator (PI)")
    for idx, co_id in enumerate(co_ids, 4):
        sub(ws.cell(2, idx), co_id)
    ws.row_dimensions[2].height = 28

    # Build PI mapping lookup
    pi_map_lookup = {}
    for m in state.pi_mappings:
        pi_map_lookup[(m.co_id, m.pi_id)] = m

    # Build PO mapping lookup
    po_map_lookup = {}
    for m in state.co_po_mapping:
        po_map_lookup[(m.co_id, m.po_id)] = m

    # Group PIs by PO and write rows
    pos_list = ["PO1", "PO2", "PO3", "PO4", "PO5", "PO6", "PO7", "PO8", "PO9", "PO10", "PO11", "PO12"]
    row_idx = 3
    
    for po_id in pos_list:
        po_pis = [pi for pi in pis if pi.po_id == po_id]
        if not po_pis:
            continue
            
        for pi in po_pis:
            ws.row_dimensions[row_idx].height = 36
            dat(ws.cell(row_idx, 1), pi.po_id, center=True)
            dat(ws.cell(row_idx, 2), pi.competency_id, center=True)
            dat(ws.cell(row_idx, 3), f"{pi.pi_id}: {pi.pi_statement}", wrap=True)
            
            for col_offset, co in enumerate(cos, 4):
                m = pi_map_lookup.get((co.co_id, pi.pi_id))
                mapped = m.mapped if m else "N"
                cell = ws.cell(row_idx, col_offset)
                dat(cell, mapped, center=True)
                if mapped == "Y":
                    cell.fill = med_fill()  # light green
                    cell.font = Font(bold=True, color="375623")
                else:
                    cell.font = Font(color="7F7F7F")
            row_idx += 1
            
        # Summary row for PO
        ws.row_dimensions[row_idx].height = 30
        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=3)
        ws.cell(row_idx, 1).value = f"PO to CO Mapping for {po_id}"
        ws.cell(row_idx, 1).font = Font(bold=True, color="FFFFFF")
        ws.cell(row_idx, 1).fill = sub_fill()
        ws.cell(row_idx, 1).alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.cell(row_idx, 1).border = thin_border()
        ws.cell(row_idx, 2).border = thin_border()
        ws.cell(row_idx, 3).border = thin_border()
        
        for col_offset, co in enumerate(cos, 4):
            m = po_map_lookup.get((co.co_id, po_id))
            strength = m.strength if m else 0
            cell = ws.cell(row_idx, col_offset)
            dat(cell, strength if strength > 0 else "-", center=True)
            cell.font = Font(bold=True)
            if strength == 3:
                cell.fill = PatternFill("solid", fgColor="C5D9F1") # light blue highlight
            elif strength == 2:
                cell.fill = PatternFill("solid", fgColor="DCE6F1")
            elif strength == 1:
                cell.fill = PatternFill("solid", fgColor="F2F5F9")
        row_idx += 1

    # Column widths
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 54
    for idx in range(4, len(cos) + 4):
        ws.column_dimensions[get_column_letter(idx)].width = 8


# ── Sheet 6B — Teaching Philosophy ───────────────────────────────────────────

def write_6B(ws, state: AgentState):
    ws.title = "6B - Teaching Philosophy"

    ws.merge_cells("A1:F1")
    hdr(ws["A1"], f"TEACHING PHILOSOPHY — {state.subject_name.upper()}", size=13)
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A3:F12")
    cell = ws["A3"]
    cell.value = state.teaching_philosophy
    cell.alignment = Alignment(wrap_text=True, vertical="top")
    cell.border = thin_border()
    ws.row_dimensions[3].height = 160

    # CO table
    ws.cell(14, 1).value = "Course Outcomes Addressed:"
    ws.cell(14, 1).font = Font(bold=True, size=11)

    sub(ws.cell(15, 1), "CO")
    sub(ws.cell(15, 2), "Statement")
    sub(ws.cell(15, 3), "Bloom's Level")
    sub(ws.cell(15, 4), "Validation Status")
    sub(ws.cell(15, 5), "Confidence")

    for i, co in enumerate(state.cos, 16):
        dat(ws.cell(i, 1), co.co_id, center=True)
        dat(ws.cell(i, 2), co.statement, wrap=True)
        dat(ws.cell(i, 3), f"L{co.blooms_level} — {co.blooms_keyword}", center=True)
        status_cell = ws.cell(i, 4)
        dat(status_cell, co.validation_status, center=True)
        if co.validation_status == "approved":
            status_cell.fill = good_fill()
        else:
            status_cell.fill = weak_fill()
        dat(ws.cell(i, 5), f"{co.confidence_score:.0%}", center=True)
        ws.row_dimensions[i].height = 32

    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 52
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 12


# ── Sheet Course Articulation Matrix ──────────────────────────────────────────

def write_articulation_matrix(ws, state: AgentState):
    ws.title = "Course Articulation Matrix"
    cos = state.cos
    pos = state.pos
    po_ids = [p.po_id for p in pos]
    mapping_lookup = {(m.co_id, m.po_id): m for m in state.co_po_mapping}

    ws.merge_cells(f"A1:{get_column_letter(len(po_ids) + 2)}1")
    hdr(ws["A1"], "COURSE ARTICULATION MATRIX (CO-PO DIRECT STRENGTHS)", size=13)
    ws.row_dimensions[1].height = 30

    headers = ["CO", "CO Statement"] + po_ids
    for col, h in enumerate(headers, 1):
        sub(ws.cell(2, col), h)
    ws.row_dimensions[2].height = 28

    # Color fills for 3, 2, 1, 0
    green_f = PatternFill("solid", fgColor="C6EFCE") # light green
    green_font = Font(bold=True, color="006100")
    
    amber_f = PatternFill("solid", fgColor="FFEB9C") # light yellow
    amber_font = Font(bold=True, color="9C6500")
    
    red_f = PatternFill("solid", fgColor="FFC7CE") # light red
    red_font = Font(bold=True, color="9C0006")
    
    gray_f = PatternFill("solid", fgColor="F2F2F2")
    gray_font = Font(color="7F7F7F")

    for row_idx, co in enumerate(cos, 3):
        ws.row_dimensions[row_idx].height = 36
        dat(ws.cell(row_idx, 1), co.co_id, center=True)
        dat(ws.cell(row_idx, 2), co.statement, wrap=True)

        for col_idx, po_id in enumerate(po_ids, 3):
            m = mapping_lookup.get((co.co_id, po_id))
            strength = m.strength if m else 0
            cell = ws.cell(row_idx, col_idx)
            
            dat(cell, strength if strength > 0 else "-", center=True)
            cell.font = Font(bold=True)
            
            if strength == 3:
                cell.fill = green_f
                cell.font = green_font
            elif strength == 2:
                cell.fill = amber_f
                cell.font = amber_font
            elif strength == 1:
                cell.fill = red_f
                cell.font = red_font
            else:
                cell.fill = gray_f
                cell.font = gray_font
                
            if m and m.reasoning:
                cell.comment = Comment(m.reasoning, "AI Mapper")

    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 54
    for i in range(3, len(po_ids) + 3):
        ws.column_dimensions[get_column_letter(i)].width = 8


# ── Sheet 6C — CO-PO Attainment Matrix ─────────────────────────────────────

def write_6C(ws, state: AgentState):
    ws.title = "6C - CO-PO Attainment Matrix"
    po_ids = [p.po_id for p in state.pos]
    mapping_lookup = {(m.co_id, m.po_id): m for m in state.co_po_mapping}
    att_lookup = {a.co_id: a for a in state.co_attainment}

    ws.merge_cells(f"A1:{get_column_letter(len(po_ids) + 4)}1")
    hdr(ws["A1"], "COURSE OUTCOME TO PO ATTAINMENT MATRIX (WEIGHTED)", size=13)
    ws.row_dimensions[1].height = 30

    headers = ["CO", "CO Statement", "Avg %", "Attainment Level"] + po_ids
    for col, h in enumerate(headers, 1):
        sub(ws.cell(2, col), h)
    ws.row_dimensions[2].height = 28

    for row_idx, co in enumerate(state.cos, 3):
        ws.row_dimensions[row_idx].height = 40
        att = att_lookup.get(co.co_id)
        level = att.achieved_level if att else 0

        dat(ws.cell(row_idx, 1), co.co_id, center=True)
        dat(ws.cell(row_idx, 2), co.statement, wrap=True)
        dat(ws.cell(row_idx, 3), f"{att.avg_percentage}%" if att else "-", center=True)

        level_cell = ws.cell(row_idx, 4)
        level_cell.value = f"Level {level}" if level > 0 else "Not Achieved"
        level_cell.alignment = Alignment(horizontal="center", vertical="center")
        level_cell.border = thin_border()
        if level == 3:   level_cell.fill = green_fill(); level_cell.font = Font(color="FFFFFF", bold=True)
        elif level == 2: level_cell.fill = med_fill()
        elif level == 1: level_cell.fill = low_fill()
        else:            level_cell.fill = red_fill()

        for col_idx, po_id in enumerate(po_ids, 5):
            m = mapping_lookup.get((co.co_id, po_id))
            strength = m.strength if m else 0
            weighted = round(level * strength / 3, 2) if strength > 0 and level > 0 else 0
            cell = ws.cell(row_idx, col_idx)
            dat(cell, weighted if weighted > 0 else "-", center=True)
            if weighted >= 2:   cell.fill = green_fill(); cell.font = Font(color="FFFFFF")
            elif weighted >= 1: cell.fill = med_fill()
            elif weighted > 0:  cell.fill = low_fill()

    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 48
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 16
    for i in range(5, len(po_ids) + 5):
        ws.column_dimensions[get_column_letter(i)].width = 9


# ── Sheet 6D — PO Attainment ──────────────────────────────────────────────────

def write_6D(ws, state: AgentState):
    ws.title = "6D - PO Attainment"

    ws.merge_cells("A1:F1")
    hdr(ws["A1"], "PROGRAM OUTCOME (PO) ATTAINMENT ANALYSIS", size=13)
    ws.row_dimensions[1].height = 30

    # Formula note
    ws.merge_cells("A3:F3")
    ws["A3"].value = "Formula: PO Attainment = Σ(CO_Attainment × Mapping_Strength) / Σ(Mapping_Strength)"
    ws["A3"].font = Font(italic=True, size=10)
    ws["A3"].alignment = Alignment(horizontal="center")

    # Headers
    headers = ["PO", "PO Statement", "Weighted Attainment", "Contributing COs", "Status", "Weakness Reason"]
    for col, h in enumerate(headers, 1):
        sub(ws.cell(5, col), h)
    ws.row_dimensions[5].height = 28

    po_lookup = {p.po_id: p for p in state.pos}

    for row_idx, att in enumerate(state.po_attainment, 6):
        ws.row_dimensions[row_idx].height = 36
        po = po_lookup.get(att.po_id)

        dat(ws.cell(row_idx, 1), att.po_id, center=True)
        dat(ws.cell(row_idx, 2), po.statement if po else "", wrap=True)
        
        score_cell = ws.cell(row_idx, 3)
        dat(score_cell, round(att.weighted_attainment, 3), center=True)
        score_cell.font = Font(bold=True)
        if att.weighted_attainment >= 2:   score_cell.fill = green_fill(); score_cell.font = Font(bold=True, color="FFFFFF")
        elif att.weighted_attainment >= 1: score_cell.fill = med_fill()
        else:                              score_cell.fill = red_fill()

        dat(ws.cell(row_idx, 4), ", ".join(att.contributing_cos), wrap=True)

        status_cell = ws.cell(row_idx, 5)
        status_cell.value = "⚠ Weak" if att.is_weak else "✓ Good"
        status_cell.alignment = Alignment(horizontal="center", vertical="center")
        status_cell.border = thin_border()
        status_cell.fill = weak_fill() if att.is_weak else good_fill()

        dat(ws.cell(row_idx, 6), att.weakness_reason or "", wrap=True)

    # Summary stats
    summary_row = len(state.po_attainment) + 8
    ws.cell(summary_row, 1).value = "Summary"
    ws.cell(summary_row, 1).font = Font(bold=True, size=11)
    weak_count = sum(1 for a in state.po_attainment if a.is_weak)
    ws.cell(summary_row + 1, 1).value = f"Total POs: {len(state.po_attainment)}"
    ws.cell(summary_row + 2, 1).value = f"Weak POs: {weak_count}"
    ws.cell(summary_row + 3, 1).value = f"Strong POs: {len(state.po_attainment) - weak_count}"
    avg = sum(a.weighted_attainment for a in state.po_attainment) / len(state.po_attainment) if state.po_attainment else 0
    ws.cell(summary_row + 4, 1).value = f"Average PO Attainment: {avg:.3f}"

    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 30


# ── Sheet 6E — Recommendations ────────────────────────────────────────────────

def write_6E(ws, state: AgentState):
    ws.title = "6E - Recommendations"

    ws.merge_cells("A1:E1")
    hdr(ws["A1"], "AI-GENERATED RECOMMENDATIONS FOR IMPROVEMENT", size=13)
    ws.row_dimensions[1].height = 30

    headers = ["Target", "Priority", "Issue Identified", "Recommendation", "Action Type"]
    for col, h in enumerate(headers, 1):
        sub(ws.cell(3, col), h)
    ws.row_dimensions[3].height = 28

    for row_idx, rec in enumerate(state.recommendations, 4):
        ws.row_dimensions[row_idx].height = 50

        dat(ws.cell(row_idx, 1), rec.target, center=True)

        priority_cell = ws.cell(row_idx, 2)
        priority_cell.value = rec.priority
        priority_cell.alignment = Alignment(horizontal="center", vertical="center")
        priority_cell.border = thin_border()
        if rec.priority == "High":   priority_cell.fill = red_fill()
        elif rec.priority == "Medium": priority_cell.fill = low_fill()
        else:                          priority_cell.fill = good_fill()

        dat(ws.cell(row_idx, 3), rec.issue, wrap=True)
        dat(ws.cell(row_idx, 4), rec.suggestion, wrap=True)

        # Auto-classify action type
        suggestion_lower = rec.suggestion.lower()
        if "lab" in suggestion_lower or "practical" in suggestion_lower:
            action = "Practical Enhancement"
        elif "assignment" in suggestion_lower or "project" in suggestion_lower:
            action = "Assessment Redesign"
        elif "lecture" in suggestion_lower or "teach" in suggestion_lower:
            action = "Pedagogy Improvement"
        else:
            action = "Curriculum Review"
        dat(ws.cell(row_idx, 5), action, center=True)

    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 35
    ws.column_dimensions["D"].width = 45
    ws.column_dimensions["E"].width = 22


# ── Sheet 6F — Audit Trail ────────────────────────────────────────────────────

def write_6F(ws, state: AgentState):
    ws.title = "6F - Audit Trail"

    ws.merge_cells("A1:D1")
    hdr(ws["A1"], "AGENT AUDIT TRAIL — EXPLAINABLE AI LOG", size=13)
    ws.row_dimensions[1].height = 30

    headers = ["Timestamp", "Agent", "Action", "Detail"]
    for col, h in enumerate(headers, 1):
        sub(ws.cell(3, col), h)

    for row_idx, entry in enumerate(state.audit_trail, 4):
        ws.row_dimensions[row_idx].height = 28
        dat(ws.cell(row_idx, 1), entry.get("timestamp", ""), center=True)
        dat(ws.cell(row_idx, 2), entry.get("agent", ""))
        dat(ws.cell(row_idx, 3), entry.get("action", ""))
        dat(ws.cell(row_idx, 4), entry.get("detail", ""), wrap=True)

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 26
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 55


def write_direct_attainment(ws, state: AgentState):
    ws.title = "Direct Attainment Matrix"
    cos = state.cos
    att_lookup = {a.co_id: a for a in state.co_attainment}
    
    # Title
    ws.merge_cells("A1:M1")
    hdr(ws["A1"], "COURSE OUTCOMES - DIRECT ATTAINMENT", size=13)
    ws.row_dimensions[1].height = 30
    
    # Column headers - Row 2
    ws.merge_cells("A2:A4")
    sub(ws["A2"], "Sr. NO.")
    ws.merge_cells("B2:B4")
    sub(ws["B2"], "CO")
    ws.merge_cells("C2:C4")
    sub(ws["C2"], "COURSE OUTCOMES")
    ws.merge_cells("D2:D4")
    sub(ws["D2"], "REVISED BLOOMS LEVEL")
    ws.merge_cells("E2:E4")
    sub(ws["E2"], "TARGET FOR COs %")
    
    ws.merge_cells("F2:I2")
    sub(ws["F2"], "CIE ASSESSMENT")
    ws.merge_cells("J2:K2")
    sub(ws["J2"], "ESE ASSESSMENT")
    
    ws.merge_cells("L2:L3")
    sub(ws["L2"], "AVERAGE ACHIEVED")
    ws.merge_cells("M2:M4")
    sub(ws["M2"], "ATTAINMENT LEVEL")
    
    # Sub-headers - Row 3
    sub(ws.cell(3, 6), "IA")
    sub(ws.cell(3, 7), "MSE")
    sub(ws.cell(3, 8), "TOTAL")
    sub(ws.cell(3, 9), "ATTAINMENT LEVEL")
    sub(ws.cell(3, 10), "ESE")
    sub(ws.cell(3, 11), "ATTAINMENT LEVEL")
    
    # Weights - Row 4
    ia_max = 30.0
    if state.ia_max_marks:
        ia_max = next(iter(state.ia_max_marks.values()), 30.0)
    mse_max = 20.0
    if state.mse_max_marks:
        mse_max = next(iter(state.mse_max_marks.values()), 20.0)
    cie_max = ia_max + mse_max
    ese_max = 50.0
    if state.ese_max_marks:
        ese_max = next(iter(state.ese_max_marks.values()), 50.0)
    total_max = cie_max + ese_max
    
    for c in range(1, 14):
        ws.cell(4, c).border = thin_border()
        ws.cell(4, c).font = Font(bold=True)
        ws.cell(4, c).alignment = Alignment(horizontal="center", vertical="center")
        
    ws.cell(4, 6).value = ia_max
    ws.cell(4, 7).value = mse_max
    ws.cell(4, 8).value = cie_max
    ws.cell(4, 10).value = ese_max
    ws.cell(4, 12).value = total_max
    
    ws.row_dimensions[2].height = 20
    ws.row_dimensions[3].height = 20
    ws.row_dimensions[4].height = 20
    
    # Write data
    row_idx = 5
    active_cos = [co for co in cos if co.statement and co.statement.strip()]
    
    # Helper to convert None to empty string
    def fmt(val):
        return round(val, 2) if val is not None else ""
        
    for idx, co in enumerate(cos):
        ws.row_dimensions[row_idx].height = 36
        att = att_lookup.get(co.co_id)
        is_active = co.statement and co.statement.strip()
        
        dat(ws.cell(row_idx, 1), idx + 1, center=True)
        dat(ws.cell(row_idx, 2), co.co_id, center=True)
        dat(ws.cell(row_idx, 3), co.statement if is_active else "", wrap=True)
        dat(ws.cell(row_idx, 4), f"Level {co.blooms_level}" if is_active else "0", center=True)
        dat(ws.cell(row_idx, 5), co.target_attainment if is_active else 0, center=True)
        
        if is_active and att:
            dat(ws.cell(row_idx, 6), fmt(att.ia_percentage), center=True)
            dat(ws.cell(row_idx, 7), fmt(att.mse_percentage), center=True)
            dat(ws.cell(row_idx, 8), fmt(att.cie_percentage), center=True)
            dat(ws.cell(row_idx, 9), fmt(att.cie_level), center=True)
            dat(ws.cell(row_idx, 10), fmt(att.ese_percentage), center=True)
            dat(ws.cell(row_idx, 11), fmt(att.ese_level), center=True)
            dat(ws.cell(row_idx, 12), fmt(att.avg_percentage), center=True)
            dat(ws.cell(row_idx, 13), fmt(att.achieved_level), center=True)
        else:
            for c in range(6, 14):
                dat(ws.cell(row_idx, c), "", center=True)
        row_idx += 1
        
    # Averages row
    ws.row_dimensions[row_idx].height = 24
    ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=4)
    avg_lbl = ws.cell(row_idx, 1)
    avg_lbl.value = "Average"
    avg_lbl.font = Font(bold=True)
    avg_lbl.alignment = Alignment(horizontal="center", vertical="center")
    avg_lbl.border = thin_border()
    for c in range(2, 5):
        ws.cell(row_idx, c).border = thin_border()
        
    # Calculate Averages
    active_atts = [att_lookup[co.co_id] for co in active_cos if co.co_id in att_lookup]
    
    def get_avg(extractor):
        vals = [extractor(a) for a in active_atts if extractor(a) is not None]
        return round(sum(vals) / len(vals), 2) if vals else ""
        
    avg_target = round(sum(co.target_attainment for co in cos) / len(cos)) if cos else 0
    
    dat(ws.cell(row_idx, 5), avg_target, center=True)
    ws.cell(row_idx, 5).font = Font(bold=True)
    
    avg_cols = {
        6: lambda a: a.ia_percentage,
        7: lambda a: a.mse_percentage,
        8: lambda a: a.cie_percentage,
        9: lambda a: a.cie_level,
        10: lambda a: a.ese_percentage,
        11: lambda a: a.ese_level,
        12: lambda a: a.avg_percentage,
        13: lambda a: a.achieved_level
    }
    
    for col, extractor in avg_cols.items():
        cell = ws.cell(row_idx, col)
        val = get_avg(extractor)
        dat(cell, val, center=True)
        cell.font = Font(bold=True)
        
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 8
    ws.column_dimensions["C"].width = 48
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 16
    for c in range(6, 14):
        ws.column_dimensions[get_column_letter(c)].width = 12

# ── Main Runner ───────────────────────────────────────────────────────────────

def run(state: AgentState) -> str:
    state.log("ReportGeneratorAgent", "start", "Writing Excel report")
    os.makedirs("data/output", exist_ok=True)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    write_6A(wb.create_sheet(), state)
    write_6B(wb.create_sheet(), state)
    write_articulation_matrix(wb.create_sheet(), state)
    write_direct_attainment(wb.create_sheet(), state)
    write_6C(wb.create_sheet(), state)
    write_6D(wb.create_sheet(), state)
    write_6E(wb.create_sheet(), state)
    write_6F(wb.create_sheet(), state)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M")
    safe_name = state.subject_name.replace(" ", "_")
    path = f"data/output/{safe_name}_MultiAgent_{timestamp}.xlsx"
    wb.save(path)

    state.log("ReportGeneratorAgent", "complete", f"Saved to {path}")
    return path